import openpyxl
import os
import pprint
def defectExcel(FileName):
    headers = {}
    severityList = {}
    statusList = {}
    total_defects = 0
    location = os.path.join('C:\\Users\\visswami\\Downloads',FileName)
    defectBook = openpyxl.load_workbook(location)#  open workbook inputted by user
    defectSheet = defectBook.get_sheet_by_name('cdets-results') # open worksheet by name
    for rowTitle in range(1,defectSheet.min_row + 1):#get all headings from excel worksheet and store in headers dictionary
        for columnTitle in range(1,defectSheet.max_column):
            headers.setdefault(defectSheet.cell(row = rowTitle, column = columnTitle).value,0)
            headers[defectSheet.cell(row = rowTitle, column = columnTitle).value] = \
                defectSheet.cell(row = rowTitle, column = columnTitle).coordinate# add title as key and and coordinate as value in headers eg, status = A11
    col_tuple = openpyxl.utils.coordinate_from_string(headers['Severity-desc'])# create a tuple for Severity coordinate and store
    col_tuple_status = openpyxl.utils.coordinate_from_string(headers['Status-desc'])# create a tuple for Status coordinate and store
    col = openpyxl.utils.column_index_from_string(col_tuple[0])#convert column value to integer value,eg B to 2 or C to 3
    col_status = openpyxl.utils.column_index_from_string(col_tuple_status[0])
    for i in range(2, defectSheet.max_row + 1):#iterate through the list for severity and status from row 2 to end of the row
        severityList.setdefault(defectSheet.cell(row = i, column = col).value,0)
        statusList.setdefault(defectSheet.cell(row = i, column = col_status).value,0)
        severityList[defectSheet.cell(row = i, column = col).value] = severityList[defectSheet.cell(row = i, column = col).value] + 1
        statusList[defectSheet.cell(row = i, column = col_status).value] = statusList[defectSheet.cell(row = i, column = col_status).value] + 1
    for total in statusList.values():#get total number of defects
        total_defects = total_defects + total
    valid_defects = [statusList.get('A-Assigned',0),statusList.get('N-New',0),statusList.get('O-Open',0),
                     statusList.get('I-Info_req',0),statusList.get('P-Postponed',0)]#get valid defects and store in list
    invalid_defects = [statusList.get('D-Duplicate',0), statusList.get('J-Junked',0),
                       statusList.get('U-Unreproducible',0),statusList.get('C-Closed',0)]# get invalid defects and store in list
    resolved_defects = [statusList.get('R-Resolved',0),statusList.get('V-Verified',0)]#get resolved defects and store in list
    print('Total defects = {}'.format(total_defects))
    print('Total valid defects = {}'.format(sum(valid_defects)))#sum all the valid defects
    print('Total Resolved defects = {}'.format(sum(resolved_defects)))#sum all resolved defects
    print('Total invalid defects = {}'.format(sum(invalid_defects)))#sum of invalid defects
    #pprint.pprint(severityList)
    #pprint.pprint(statusList)
if __name__ == '__main__':
    defectExcel('cdets-results.xlsx')