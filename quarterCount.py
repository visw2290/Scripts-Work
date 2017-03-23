import openpyxl
import os
import pprint
import datetime
def quarterExcel(FileName):
    headers = {}
    severityList = {}
    statusList = {}
    location = os.path.join('C:\\Users\\visswami\\Downloads',FileName)
    defectBook = openpyxl.load_workbook(location)
    defectSheet = defectBook.get_sheet_by_name('cdets-results')
    for rowTitle in range(1,defectSheet.min_row + 1):
        for columnTitle in range(1,defectSheet.max_column + 1):
            headers.setdefault(defectSheet.cell(row = rowTitle, column = columnTitle).value,0)
            headers[defectSheet.cell(row = rowTitle, column = columnTitle).value] = defectSheet.cell(row = rowTitle, column = columnTitle).coordinate
    col_tuple = openpyxl.utils.coordinate_from_string(headers['Severity-desc'])
    col_tuple_status = openpyxl.utils.coordinate_from_string(headers['Status-desc'])
    col = openpyxl.utils.column_index_from_string(col_tuple[0])
    col_status = openpyxl.utils.column_index_from_string(col_tuple_status[0])
    for i in range(2, defectSheet.max_row + 1):
        if defectSheet.cell(row = i, column = defectSheet.max_column).value >= datetime.datetime(2016,11,1):#iterate only when date is greater than certain value
            severityList.setdefault(defectSheet.cell(row = i, column = col).value,0)
            statusList.setdefault(defectSheet.cell(row = i, column = col_status).value,0)
            severityList[defectSheet.cell(row = i, column = col).value] = severityList[defectSheet.cell(row = i, column = col).value] + 1
            statusList[defectSheet.cell(row=i, column=col_status).value] = statusList[defectSheet.cell(row=i,column=col_status).value] + 1
    total_defects = 0
    for total in statusList.values():
        total_defects = total_defects + total
    valid_defects = [statusList.get('A-Assigned',0),statusList.get('N-New',0),statusList.get('O-Open',0),
                     statusList.get('I-Info_req',0),statusList.get('P-Postponed',0)]
    invalid_defects = [statusList.get('D-Duplicate',0), statusList.get('J-Junked',0),statusList.get('U-Unreproducible',0),statusList.get('C-Closed',0)]
    resolved_defects = [statusList.get('R-Resolved',0),statusList.get('V-Verified',0)]
    print('Total Defects = {}'.format(total_defects))
    print('Total Valid defects = {}'.format(sum(valid_defects)))
    print('Total Resolved defects = {}'.format(sum(resolved_defects)))
    print('Total Invalid defects = {}'.format(sum(invalid_defects)))
    #pprint.pprint(severityList)
    #pprint.pprint(statusList)
if __name__ == '__main__':
    quarterExcel('cdets-results.xlsx')